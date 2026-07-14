# pages/estimate.py - Смета на работы

import streamlit as st
import pandas as pd
from datetime import datetime
import io
import json
import os

# Путь для сохранения смет
ESTIMATES_FILE = 'data/estimates.json'

def save_estimate_to_file(data):
    """Сохранение сметы в JSON файл"""
    try:
        os.makedirs(os.path.dirname(ESTIMATES_FILE), exist_ok=True)
        
        # Загружаем существующие сметы
        if os.path.exists(ESTIMATES_FILE):
            with open(ESTIMATES_FILE, 'r', encoding='utf-8') as f:
                estimates = json.load(f)
        else:
            estimates = []
        
        estimates.append(data)
        
        with open(ESTIMATES_FILE, 'w', encoding='utf-8') as f:
            json.dump(estimates, f, ensure_ascii=False, indent=2, default=str)
        return True
    except Exception as e:
        st.error(f"Ошибка сохранения: {e}")
        return False

def load_estimates_from_file():
    """Загрузка смет из файла"""
    try:
        if os.path.exists(ESTIMATES_FILE):
            with open(ESTIMATES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return []
    except Exception as e:
        return []

def calculate_cost(quantity, price, discount=0, vat=0):
    """Правильный расчет стоимости"""
    base = quantity * price
    discount_amount = base * (discount / 100)
    after_discount = base - discount_amount
    vat_amount = after_discount * (vat / 100)
    total = after_discount + vat_amount
    
    return {
        'base': base,
        'discount_amount': discount_amount,
        'after_discount': after_discount,
        'vat_amount': vat_amount,
        'total': total
    }

def show_estimate():
    st.title("📋 Смета на работы")
    st.markdown("---")
    
    # --- ИНФОРМАЦИЯ О СМЕТЕ ---
    col1, col2 = st.columns(2)
    with col1:
        estimate_name = st.text_input("Название сметы", 
                                     f"Смета №{len(load_estimates_from_file()) + 1}")
    
    with col2:
        estimate_date = st.date_input("Дата", datetime.now())
    
    col1, col2 = st.columns(2)
    with col1:
        client_options = [c['name'] for c in st.session_state.clients] + ["➕ Новый клиент"]
        selected_client = st.selectbox("Клиент", client_options)
        
        if selected_client == "➕ Новый клиент":
            with st.expander("Добавить клиента"):
                new_name = st.text_input("Имя")
                new_phone = st.text_input("Телефон")
                if st.button("➕ Добавить клиента"):
                    if new_name:
                        st.session_state.clients.append({
                            'id': len(st.session_state.clients) + 1,
                            'name': new_name,
                            'phone': new_phone
                        })
                        st.success(f"✅ Клиент {new_name} добавлен")
                        st.rerun()
        else:
            estimate_client = selected_client
    
    with col2:
        object_options = [o['name'] for o in st.session_state.objects] + ["➕ Новый объект"]
        selected_object = st.selectbox("Объект (адрес)", object_options)
        
        if selected_object == "➕ Новый объект":
            with st.expander("Добавить объект"):
                new_obj_name = st.text_input("Название")
                new_obj_address = st.text_input("Адрес")
                if st.button("➕ Добавить объект"):
                    if new_obj_name:
                        st.session_state.objects.append({
                            'id': len(st.session_state.objects) + 1,
                            'name': new_obj_name,
                            'address': new_obj_address
                        })
                        st.success(f"✅ Объект {new_obj_name} добавлен")
                        st.rerun()
        else:
            estimate_object = selected_object
    
    st.markdown("---")
    
    # --- ДОБАВЛЕНИЕ РАБОТЫ ---
    st.subheader("➕ Добавить позицию в смету")
    
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        room_name = st.text_input("Помещение", placeholder="Коридор, Зал, Кухня...")
    
    with col2:
        work_types = list(st.session_state.price_list.keys())
        selected_work_type = st.selectbox("Тип работ", work_types)
    
    with col3:
        if selected_work_type:
            categories = list(st.session_state.price_list[selected_work_type].keys())
            selected_category = st.selectbox("Категория", categories)
        else:
            selected_category = None
    
    if selected_work_type and selected_category:
        works = st.session_state.price_list[selected_work_type].get(selected_category, {})
        if works:
            work_name = st.selectbox("Услуга", list(works.keys()))
            work_info = works[work_name]
        else:
            work_name = None
            work_info = None
            st.warning("Нет работ в этой категории")
    else:
        work_name = None
        work_info = None
    
    if work_info:
        col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
        
        with col1:
            unit = work_info['unit']
            quantity = st.number_input(f"Кол-во ({unit})", min_value=0.0, step=0.5, value=1.0)
        
        with col2:
            price = work_info['price']
            st.write(f"**Цена:** {price:.2f} руб.")
        
        with col3:
            discount = st.number_input("Скидка %", min_value=0.0, max_value=100.0, value=0.0, step=1.0)
        
        with col4:
            vat = st.number_input("НДС %", min_value=0.0, max_value=100.0, value=20.0, step=1.0)
        
        if quantity > 0:
            cost_data = calculate_cost(quantity, price, discount, vat)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("💰 Базовая стоимость", f"{cost_data['base']:,.2f} руб.")
            with col2:
                st.metric("📉 Скидка", f"{cost_data['discount_amount']:,.2f} руб.")
            with col3:
                st.metric("📈 НДС", f"{cost_data['vat_amount']:,.2f} руб.")
            with col4:
                st.metric("✅ Итого", f"{cost_data['total']:,.2f} руб.")
            
            if st.button("➕ Добавить в смету", key="add_work_estimate"):
                if room_name:
                    st.session_state.current_estimate.append({
                        'room': room_name,
                        'work_type': selected_work_type,
                        'category': selected_category,
                        'work': work_name,
                        'unit': unit,
                        'quantity': quantity,
                        'price': price,
                        'discount': discount,
                        'vat': vat,
                        'base_cost': cost_data['base'],
                        'discount_amount': cost_data['discount_amount'],
                        'vat_amount': cost_data['vat_amount'],
                        'total': cost_data['total']
                    })
                    st.success(f"✅ Добавлено: {room_name} → {work_name}")
                    st.rerun()
                else:
                    st.error("Введите название помещения")
    
    st.markdown("---")
    
    # --- ОТОБРАЖЕНИЕ СМЕТЫ ---
    st.subheader("📄 Смета")
    
    if st.session_state.current_estimate:
        df = pd.DataFrame(st.session_state.current_estimate)
        
        display_cols = ['room', 'category', 'work_type', 'work', 'unit', 'quantity', 'price', 'discount', 'vat', 'total']
        column_config = {
            'room': 'Помещение',
            'category': 'Категория',
            'work_type': 'Тип работ',
            'work': 'Услуга',
            'unit': 'Ед. изм.',
            'quantity': 'Кол-во',
            'price': 'Цена',
            'discount': 'Скидка %',
            'vat': 'НДС %',
            'total': 'Итого'
        }
        
        st.dataframe(
            df[display_cols],
            column_config=column_config,
            hide_index=True,
            use_container_width=True
        )
        
        # --- ИТОГИ ---
        st.markdown("---")
        st.subheader("📊 Итоги")
        
        total_cost = df['total'].sum()
        total_discount = df['discount_amount'].sum()
        total_vat = df['vat_amount'].sum()
        total_base = df['base_cost'].sum()
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("💰 Базовая стоимость", f"{total_base:,.2f} руб.")
        with col2:
            st.metric("📉 Скидка", f"{total_discount:,.2f} руб.")
        with col3:
            st.metric("📈 НДС", f"{total_vat:,.2f} руб.")
        with col4:
            st.metric("✅ ИТОГО", f"{total_cost:,.2f} руб.")
        
        # --- СВОД ПО ПОМЕЩЕНИЯМ ---
        st.markdown("---")
        st.subheader("🏠 Свод по помещениям")
        
        room_summary = df.groupby('room').agg({
            'total': 'sum',
            'quantity': 'sum'
        }).reset_index()
        room_summary.columns = ['Помещение', 'Стоимость', 'Кол-во работ']
        room_summary['Стоимость'] = room_summary['Стоимость'].apply(lambda x: f"{x:,.2f}")
        
        st.dataframe(room_summary, use_container_width=True, hide_index=True)
        
        # --- КНОПКИ ---
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("💾 Сохранить смету", key="save_estimate_full"):
                if estimate_name and estimate_client and estimate_object:
                    estimate_data = {
                        'name': estimate_name,
                        'date': estimate_date.strftime("%Y-%m-%d"),
                        'client': estimate_client,
                        'object': estimate_object,
                        'items': st.session_state.current_estimate.copy(),
                        'total': total_cost,
                        'discount_total': total_discount,
                        'vat_total': total_vat,
                        'base_total': total_base
                    }
                    
                    if save_estimate_to_file(estimate_data):
                        st.session_state.current_estimate = []
                        st.success("✅ Смета сохранена!")
                        st.rerun()
                    else:
                        st.error("Ошибка сохранения сметы")
                else:
                    st.error("Заполните название сметы, клиента и объект")
        
        with col2:
            if st.button("🗑️ Очистить смету", key="clear_estimate_full"):
                st.session_state.current_estimate = []
                st.rerun()
        
        with col3:
            if st.button("📊 Экспорт в Excel", key="export_excel"):
                export_to_excel(df, estimate_name, total_cost, total_discount, total_vat, total_base)
    else:
        st.info("Смета пуста. Добавьте работы выше.")


def export_to_excel(df, filename, total, total_discount, total_vat, total_base):
    """Экспорт в Excel"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Лист сметы
        export_df = df[['room', 'category', 'work_type', 'work', 'unit', 'quantity', 'price', 'discount', 'total']].copy()
        export_df.columns = ['Помещение', 'Категория', 'Тип работ', 'Услуга', 'ед.изм.', 'Кол-во', 'Цена', 'Скидка', 'Стоимость']
        export_df.to_excel(writer, sheet_name='Смета', index=False)
        
        # Настройка ширины
        worksheet = writer.sheets['Смета']
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 35
        worksheet.column_dimensions['E'].width = 10
        worksheet.column_dimensions['F'].width = 10
        worksheet.column_dimensions['G'].width = 12
        worksheet.column_dimensions['H'].width = 10
        worksheet.column_dimensions['I'].width = 15
        
        # Итоги
        from openpyxl.styles import Font
        last_row = len(export_df) + 3
        
        worksheet.cell(row=last_row, column=9, value=f"Базовая стоимость: {total_base:,.2f} руб.")
        worksheet.cell(row=last_row, column=9).font = Font(bold=True)
        worksheet.cell(row=last_row+1, column=9, value=f"Скидка: {total_discount:,.2f} руб.")
        worksheet.cell(row=last_row+1, column=9).font = Font(bold=True)
        worksheet.cell(row=last_row+2, column=9, value=f"НДС: {total_vat:,.2f} руб.")
        worksheet.cell(row=last_row+2, column=9).font = Font(bold=True)
        worksheet.cell(row=last_row+3, column=9, value=f"ИТОГО: {total:,.2f} руб.")
        worksheet.cell(row=last_row+3, column=9).font = Font(bold=True)
    
    st.download_button(
        label="📥 Скачать Excel",
        data=output.getvalue(),
        file_name=f"{filename}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

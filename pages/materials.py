# pages/materials.py - Смета на материалы (как в Excel)

import streamlit as st
import pandas as pd
from datetime import datetime
from utils.data_manager import save_materials_estimate, load_materials_estimates
import io

def show_materials():
    st.title("🧾 Смета на материалы")
    st.markdown("---")
    
    # --- ИНФОРМАЦИЯ О СМЕТЕ ---
    col1, col2 = st.columns(2)
    with col1:
        materials_name = st.text_input("Название сметы на материалы", 
                                      f"Смета материалов №{len(load_materials_estimates()) + 1}")
    
    with col2:
        materials_date = st.date_input("Дата", datetime.now())
    
    # Клиент и объект
    col1, col2 = st.columns(2)
    with col1:
        client_options = [c['name'] for c in st.session_state.clients] + ["➕ Новый клиент"]
        selected_client = st.selectbox("Клиент", client_options, key="mat_client")
        
        if selected_client == "➕ Новый клиент":
            with st.expander("Добавить клиента"):
                new_name = st.text_input("Имя", key="mat_new_name")
                new_phone = st.text_input("Телефон", key="mat_new_phone")
                if st.button("➕ Добавить клиента", key="mat_add_client"):
                    if new_name:
                        st.session_state.clients.append({
                            'id': len(st.session_state.clients) + 1,
                            'name': new_name,
                            'phone': new_phone
                        })
                        st.success(f"✅ Клиент {new_name} добавлен")
                        st.rerun()
        else:
            estimate_client = selected_client
    
    with col2:
        object_options = [o['name'] for o in st.session_state.objects] + ["➕ Новый объект"]
        selected_object = st.selectbox("Объект (адрес)", object_options, key="mat_object")
        
        if selected_object == "➕ Новый объект":
            with st.expander("Добавить объект"):
                new_obj_name = st.text_input("Название", key="mat_new_obj_name")
                new_obj_address = st.text_input("Адрес", key="mat_new_obj_address")
                if st.button("➕ Добавить объект", key="mat_add_obj"):
                    if new_obj_name:
                        st.session_state.objects.append({
                            'id': len(st.session_state.objects) + 1,
                            'name': new_obj_name,
                            'address': new_obj_address
                        })
                        st.success(f"✅ Объект {new_obj_name} добавлен")
                        st.rerun()
        else:
            estimate_object = selected_object
    
    st.markdown("---")
    
    # --- ДОБАВЛЕНИЕ МАТЕРИАЛА ---
    st.subheader("➕ Добавить материал")
    
    col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
    
    with col1:
        name = st.text_input("Наименование материала", key="mat_name")
        container = st.text_input("Тара (упаковка)", placeholder="мешок, ведро, рулон...", key="mat_container")
    
    with col2:
        unit = st.text_input("Ед. изм.", placeholder="шт, м², кг...", key="mat_unit")
    
    with col3:
        qty = st.number_input("Кол-во", min_value=0.0, step=1.0, key="mat_qty")
    
    with col4:
        price = st.number_input("Цена (руб.)", min_value=0.0, step=10.0, key="mat_price")
    
    if st.button("➕ Добавить материал", key="add_mat"):
        if name and unit and qty > 0 and price > 0:
            if 'estimate_materials' not in st.session_state:
                st.session_state.estimate_materials = []
            
            st.session_state.estimate_materials.append({
                '№': len(st.session_state.estimate_materials) + 1,
                'name': name,
                'container': container if container else "-",
                'unit': unit,
                'quantity': qty,
                'price': price,
                'cost': qty * price
            })
            st.success(f"✅ Материал '{name}' добавлен!")
            st.rerun()
        else:
            st.error("Заполните все поля (наименование, ед.изм., кол-во, цена)")
    
    st.markdown("---")
    
    # --- ОТОБРАЖЕНИЕ МАТЕРИАЛОВ ---
    st.subheader("📄 Смета на материалы")
    
    if 'estimate_materials' in st.session_state and st.session_state.estimate_materials:
        df = pd.DataFrame(st.session_state.estimate_materials)
        
        # Обновляем номера
        df['№'] = range(1, len(df) + 1)
        
        # Колонки как в Excel (ваш файл)
        display_cols = ['№', 'name', 'container', 'unit', 'quantity', 'price', 'cost']
        column_config = {
            '№': '№',
            'name': 'Наименование материала',
            'container': 'Тара',
            'unit': 'Ед-ца измер.',
            'quantity': 'Кол-во',
            'price': 'Цена, руб.',
            'cost': 'Сумма, руб.'
        }
        
        st.dataframe(
            df[display_cols],
            column_config=column_config,
            hide_index=True,
            use_container_width=True
        )
        
        # --- ИТОГ ---
        st.markdown("---")
        total_cost = df['cost'].sum()
        
        # Строка с итогом как в Excel
        st.markdown(f"### **ИТОГО: {total_cost:,.2f} руб.**")
        
        # --- КНОПКИ ---
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("💾 Сохранить смету", key="save_materials"):
                if materials_name and estimate_client and estimate_object:
                    materials_data = {
                        'name': materials_name,
                        'date': materials_date.strftime("%Y-%m-%d"),
                        'client': estimate_client,
                        'object': estimate_object,
                        'items': st.session_state.estimate_materials.copy(),
                        'total': total_cost
                    }
                    
                    if save_materials_estimate(materials_data):
                        st.session_state.estimate_materials = []
                        st.success("✅ Смета материалов сохранена!")
                        st.rerun()
                    else:
                        st.error("Ошибка сохранения")
                else:
                    st.error("Заполните название сметы, клиента и объект")
        
        with col2:
            if st.button("🗑️ Очистить", key="clear_materials"):
                st.session_state.estimate_materials = []
                st.rerun()
        
        with col3:
            # ЭКСПОРТ В EXCEL КАК В ВАШЕМ ФАЙЛЕ
            if st.button("📊 Экспорт в Excel", key="export_materials_excel"):
                export_materials_to_excel(df, materials_name, estimate_client, estimate_object, total_cost)
    else:
        st.info("Нет добавленных материалов")

def export_materials_to_excel(df, filename, client, object_name, total):
    """Экспорт материалов в Excel как в вашем файле"""
    
    # Создаем Excel файл в памяти
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Копируем данные для экспорта
        export_df = df[['№', 'name', 'container', 'unit', 'quantity', 'price', 'cost']].copy()
        export_df.columns = ['№', 'Наименование материала', 'Тара', 'Ед-ца измер.', 'Кол-во', 'Цена, руб.', 'Сумма, руб.']
        
        # Записываем данные
        export_df.to_excel(writer, sheet_name='Материалы', index=False)
        
        # Настройка ширины колонок (как в вашем файле)
        worksheet = writer.sheets['Материалы']
        worksheet.column_dimensions['A'].width = 8   # №
        worksheet.column_dimensions['B'].width = 35  # Наименование
        worksheet.column_dimensions['C'].width = 15  # Тара
        worksheet.column_dimensions['D'].width = 12  # Ед-ца измер.
        worksheet.column_dimensions['E'].width = 10  # Кол-во
        worksheet.column_dimensions['F'].width = 15  # Цена
        worksheet.column_dimensions['G'].width = 15  # Сумма
        
        # Добавляем строку с итогом
        from openpyxl.utils import get_column_letter
        last_row = len(export_df) + 2
        worksheet.cell(row=last_row, column=7, value=f"ИТОГО: {total:,.2f} руб.")
        
        # Жирный шрифт для итога
        from openpyxl.styles import Font
        worksheet.cell(row=last_row, column=7).font = Font(bold=True)
    
    # Скачивание
    st.download_button(
        label="📥 Скачать Excel",
        data=output.getvalue(),
        file_name=f"{filename}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
